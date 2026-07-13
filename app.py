from flask import (
    Flask,
    render_template,
    request,
    redirect,
    url_for,
    session,
    flash
)

from config import Config
from excel_reader import get_report_sheets
try:

    from excel_to_image import export_monthly_reports
except Exception:
    export_monthly_reports = None
import os

app = Flask(__name__)
app.config.from_object(Config)

UPLOAD_FOLDER = app.config["UPLOAD_FOLDER"]

ALLOWED_EXTENSIONS = {"xlsx"}

os.makedirs(UPLOAD_FOLDER, exist_ok=True)


# -------------------------------------------------
# TEMP REPORTS
# (Will be replaced by Excel in Milestone 4)
# -------------------------------------------------
def allowed_file(filename):
    return (
        "." in filename
        and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS
    )

# -------------------------------------------------
# HOME
# -------------------------------------------------
from openpyxl import load_workbook


def get_sheet_data(sheet_name):
    """
    Return worksheet data as a list of rows.
    """

    workbook = load_workbook(
        os.path.join(app.config["UPLOAD_FOLDER"], "monthly_report.xlsx"),
        data_only=True
    )

    if sheet_name not in workbook.sheetnames:
        return []

    sheet = workbook[sheet_name]

    data = []

    for row in sheet.iter_rows(values_only=True):
        data.append(row)

    return data

@app.route("/")
@app.route("/report/<sheet_name>")
def home(sheet_name=None):

    reports = get_report_sheets()

    if not reports:
        return render_template(
            "index.html",
            reports=[],
            selected_sheet=None,
            sheet_data=[]
        )

    if sheet_name is None:
        sheet_name = reports[0]

    report_image = (
        sheet_name.replace("/", "-")
                  .replace("\\", "-")
                  .replace(":", "")
                  + ".png"
    )

    return render_template(
        "index.html",
        reports=reports,
        selected_sheet=sheet_name,
        report_image=report_image
    )

# -------------------------------------------------
# LOGIN
# -------------------------------------------------
@app.route("/login", methods=["GET", "POST"])
def login():

    # Already logged in
    if session.get("logged_in"):
        return redirect(url_for("dashboard"))

    if request.method == "POST":

        email = request.form.get("email", "").strip()
        password = request.form.get("password", "")

        if (
            email == app.config["ADMIN_EMAIL"]
            and password == app.config["ADMIN_PASSWORD"]
        ):

            session["logged_in"] = True
            session["admin_email"] = email

            flash(
                "Login successful.",
                "success"
            )

            return redirect(url_for("dashboard"))

        flash(
            "Invalid email or password.",
            "danger"
        )

    return render_template("login.html")


# -------------------------------------------------
# DASHBOARD
# -------------------------------------------------
@app.route("/dashboard")
def dashboard():

    if not session.get("logged_in"):

        flash(
            "Please login first.",
            "warning"
        )

        return redirect(url_for("login"))

    reports = get_report_sheets()

    return render_template(
        "dashboard.html",
        reports=reports
    )


# -------------------------------------------------
# UPLOAD WORKBOOK
# -------------------------------------------------
@app.route("/upload", methods=["POST"])
def upload_workbook():

    if not session.get("logged_in"):
        flash("Please login first.", "warning")
        return redirect(url_for("login"))

    if "workbook" not in request.files:
        flash("Please choose an Excel workbook.", "danger")
        return redirect(url_for("dashboard"))

    file = request.files["workbook"]

    if file.filename == "":
        flash("Please choose an Excel workbook.", "danger")
        return redirect(url_for("dashboard"))

    if not allowed_file(file.filename):
        flash("Only .xlsx files are allowed.", "danger")
        return redirect(url_for("dashboard"))

    filename = "monthly_report.xlsx"

    save_path = os.path.join(
        app.config["UPLOAD_FOLDER"],
        filename
    )

    file.save(save_path)

    try:

        if export_monthly_reports:
            export_monthly_reports()

            flash(
                "Workbook uploaded and images generated successfully.",
                "success"
            )

        else:

            flash(
                "Workbook uploaded successfully.",
                "success"
            )

    except Exception:

        flash(
            "Workbook uploaded successfully. Image generation is not available on this server.",
            "warning"
        )

    return redirect(url_for("dashboard"))

# -------------------------------------------------
# LOGOUT
# -------------------------------------------------
@app.route("/logout")
def logout():

    session.clear()

    flash(
        "Logged out successfully.",
        "success"
    )

    return redirect(url_for("home"))

# -------------------------------------------------
# START
# -------------------------------------------------
if __name__ == "__main__":
    app.run(debug=True)