import os
from openpyxl import load_workbook

UPLOAD_FOLDER = "uploads"
WORKBOOK_NAME = "monthly_report.xlsx"


def workbook_path():
    """Return full workbook path."""
    return os.path.join(UPLOAD_FOLDER, WORKBOOK_NAME)


def workbook_exists():
    """Check if workbook exists."""
    return os.path.exists(workbook_path())


def get_report_sheets():
    """Return sheets to display on the website."""

    if not workbook_exists():
        return []

    wb = load_workbook(
        workbook_path(),
        data_only=True
    )

    reports = []

    for sheet in wb.sheetnames:

        name = sheet.strip()

        if name.lower() == "residents":
            continue

        if "ongoing" in name.lower():
            continue

        reports.append(name)

    # Keep "Total Pending Final" first
    if "Total Pending Final" in reports:
        reports.remove("Total Pending Final")
        reports.insert(0, "Total Pending Final")

    return reports